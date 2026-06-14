"""Reusable Excel (.xlsx) export service.

Generic, presentation-agnostic helpers for turning row data into an .xlsx file
on disk. The caller owns localization (it passes already-localized headers and a
sheet title) and is responsible for deleting the returned temp file after use.

This module is INTENTIONALLY free of subscription / Stripe / entitlement logic.
`build_user_export_rows` only reads display fields off User rows (and the pure
`user_type_admin_label` helper) — it never mutates state or evaluates access.
"""

from __future__ import annotations

import logging
import os
import tempfile
from typing import Any, Iterable, List, Optional, Sequence

logger = logging.getLogger(__name__)

# Stable column order for user exports. The headers are supplied (localized) by
# the caller in this same order; the value extractor below mirrors it.
USER_EXPORT_FIELDS = (
    "telegram_id",
    "username",
    "full_name",
    "status",
    "subscription_status",
    "case_type",
    "created_date",
    "approval_date",
)


def _fmt_dt(value: Any) -> str:
    """Render a datetime as 'YYYY-MM-DD HH:MM'; '' when missing."""
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d %H:%M")
    return str(value)


def build_user_export_rows(users: Iterable[Any]) -> List[list]:
    """Map User ORM rows to export rows (data only, no formatting/translation).

    Column order matches USER_EXPORT_FIELDS:
      Telegram User ID, Username, Full Name, Status, Subscription Status,
      Case Type, Created Date, Approval Date.
    """
    from services.user_segment import user_type_admin_label

    rows: List[list] = []
    for u in users:
        sub = getattr(u, "subscription", None)
        sub_status = str(sub.status) if sub is not None and getattr(sub, "status", None) is not None else ""
        rows.append(
            [
                getattr(u, "telegram_id", ""),
                getattr(u, "username", None) or "",
                getattr(u, "first_name", None) or "",
                getattr(u, "status", None) or "",
                sub_status,
                user_type_admin_label(u),
                _fmt_dt(getattr(u, "created_at", None)),
                _fmt_dt(getattr(u, "approved_at", None)),
            ]
        )
    return rows


# Stable column order for question exports (mirrors the headers the caller
# supplies, localized, in this same order).
QUESTION_EXPORT_FIELDS = (
    "question_id",
    "telegram_id",
    "username",
    "full_name",
    "status",
    "question_type",
    "question_text",
    "admin_reply",
    "created_date",
    "answered_date",
)


def build_question_export_rows(questions: Iterable[Any], user_map: Optional[dict] = None) -> List[list]:
    """Map Question ORM rows to export rows (data only).

    `user_map` is an optional {telegram_id: User} lookup used to fill the
    Username / Full Name columns; missing users leave those blank. Column order
    matches QUESTION_EXPORT_FIELDS.
    """
    user_map = user_map or {}
    rows: List[list] = []
    for q in questions:
        u = user_map.get(getattr(q, "user_id", None))
        rows.append(
            [
                getattr(q, "id", ""),
                getattr(q, "user_id", ""),
                (getattr(u, "username", None) or "") if u is not None else "",
                (getattr(u, "first_name", None) or "") if u is not None else "",
                getattr(q, "status", None) or "",
                getattr(q, "question_type", None) or "",
                getattr(q, "question_text", None) or "",
                getattr(q, "admin_reply_text", None) or "",
                _fmt_dt(getattr(q, "created_at", None)),
                _fmt_dt(getattr(q, "answered_at", None)),
            ]
        )
    return rows


def write_xlsx(
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    *,
    sheet_title: str = "Export",
    filename_prefix: str = "export",
) -> str:
    """Write headers + rows to a temporary .xlsx file and return its path.

    The caller MUST delete the returned file when done. A bold header row and
    best-effort column widths are applied for readability.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    # Excel caps sheet titles at 31 chars and forbids some characters.
    ws.title = (sheet_title or "Export")[:31]

    header_list = list(headers)
    ws.append(header_list)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    widths = [len(str(h)) for h in header_list]
    for row in rows:
        values = list(row)
        ws.append(values)
        for i, value in enumerate(values):
            if i < len(widths):
                widths[i] = max(widths[i], len(str(value)) if value is not None else 0)

    for i, width in enumerate(widths, start=1):
        # Pad a little, and keep columns within a sane bound.
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 60)

    fd, path = tempfile.mkstemp(prefix=f"{filename_prefix}_", suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    return path


def export_users_xlsx(
    users: Iterable[Any],
    headers: Sequence[str],
    *,
    sheet_title: str = "Users",
    filename_prefix: str = "users_export",
) -> str:
    """Convenience wrapper: build user rows and write them to an .xlsx temp file.

    Returns the temp file path (caller deletes it).
    """
    return write_xlsx(
        headers,
        build_user_export_rows(users),
        sheet_title=sheet_title,
        filename_prefix=filename_prefix,
    )


def export_questions_xlsx(
    questions: Iterable[Any],
    headers: Sequence[str],
    *,
    user_map: Optional[dict] = None,
    sheet_title: str = "Questions",
    filename_prefix: str = "questions_export",
) -> str:
    """Convenience wrapper: build question rows and write them to an .xlsx temp file.

    Returns the temp file path (caller deletes it).
    """
    return write_xlsx(
        headers,
        build_question_export_rows(questions, user_map),
        sheet_title=sheet_title,
        filename_prefix=filename_prefix,
    )


def safe_remove(path: Optional[str]) -> None:
    """Best-effort delete of a temp export file; never raises."""
    if not path:
        return
    try:
        os.remove(path)
    except OSError as e:
        logger.warning("excel_export temp cleanup failed path=%s err=%s", path, e)
